"""Admin panel routes: dashboard, registrations, export, delete."""
import csv
import io
from datetime import datetime

from flask import (Blueprint, Response, current_app, flash, redirect,
                   render_template, request, session, url_for)

from app import db
from app.models import Participant

admin_bp = Blueprint('admin', __name__)


# ---------------------------------------------------------------------------
# Auth guard
# ---------------------------------------------------------------------------

@admin_bp.before_request
def require_admin_login():
    """Redirect to login if not authenticated, except for the login route."""
    if request.endpoint not in ('admin.login',) and not session.get('admin_logged_in'):
        return redirect(url_for('admin.login'))


# ---------------------------------------------------------------------------
# Auth routes
# ---------------------------------------------------------------------------

@admin_bp.route('/login', methods=['GET', 'POST'])
def login():
    """Admin login page."""
    if session.get('admin_logged_in'):
        return redirect(url_for('admin.dashboard'))

    if request.method == 'POST':
        password = request.form.get('password', '')
        correct = current_app.config.get('ADMIN_PASSWORD', 'admin123')
        if password == correct:
            session['admin_logged_in'] = True
            flash('Welcome back, Admin!', 'success')
            return redirect(url_for('admin.dashboard'))
        flash('Invalid password. Please try again.', 'danger')

    return render_template('admin/login.html')


@admin_bp.route('/logout')
def logout():
    """Admin logout."""
    session.pop('admin_logged_in', None)
    flash('You have been logged out.', 'info')
    return redirect(url_for('main.index'))


# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------

@admin_bp.route('/')
def dashboard():
    """Admin dashboard with statistics."""
    total = Participant.query.count()
    seminars = current_app.config['SEMINARS']
    seminar_stats = [
        {'name': s, 'count': Participant.query.filter_by(seminar=s).count()}
        for s in seminars
    ]
    recent = (Participant.query
              .order_by(Participant.timestamp.desc())
              .limit(5).all())
    return render_template('admin/dashboard.html', total=total,
                           seminar_stats=seminar_stats, recent=recent)


# ---------------------------------------------------------------------------
# Registrations list
# ---------------------------------------------------------------------------

@admin_bp.route('/registrations')
def registrations():
    """Paginated registrations list with search and filter."""
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '').strip()
    seminar_filter = request.args.get('seminar', '').strip()

    query = Participant.query

    if search:
        like = f'%{search}%'
        query = query.filter(
            db.or_(
                Participant.name.ilike(like),
                Participant.email.ilike(like),
                Participant.registration_id.ilike(like),
                Participant.organization.ilike(like),
                Participant.phone.ilike(like),
            )
        )

    if seminar_filter:
        query = query.filter_by(seminar=seminar_filter)

    pagination = (query
                  .order_by(Participant.timestamp.desc())
                  .paginate(page=page, per_page=20, error_out=False))

    return render_template(
        'admin/registrations.html',
        participants=pagination,
        search=search,
        seminar_filter=seminar_filter,
        seminars=current_app.config['SEMINARS'],
    )


# ---------------------------------------------------------------------------
# Export routes
# ---------------------------------------------------------------------------

@admin_bp.route('/export/csv')
def export_csv():
    """Export all registrations as CSV."""
    rows = Participant.query.order_by(Participant.timestamp.desc()).all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Registration ID', 'Name', 'Email', 'Phone',
                     'Organization', 'Seminar', 'Registered At'])
    for p in rows:
        writer.writerow([
            p.registration_id, p.name, p.email, p.phone,
            p.organization, p.seminar,
            p.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
        ])

    filename = f"registrations_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


@admin_bp.route('/export/excel')
def export_excel():
    """Export all registrations as Excel (.xlsx)."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        flash('openpyxl is not installed. Use CSV export instead.', 'warning')
        return redirect(url_for('admin.registrations'))

    rows = Participant.query.order_by(Participant.timestamp.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Registrations'

    headers = ['Registration ID', 'Name', 'Email', 'Phone',
               'Organization', 'Seminar', 'Registered At']
    hdr_fill = PatternFill(start_color='1F4E79', end_color='1F4E79',
                           fill_type='solid')
    hdr_font = Font(color='FFFFFF', bold=True)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center')

    for row_idx, p in enumerate(rows, 2):
        values = [p.registration_id, p.name, p.email, p.phone,
                  p.organization, p.seminar,
                  p.timestamp.strftime('%Y-%m-%d %H:%M:%S')]
        for col_idx, val in enumerate(values, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Auto-size columns
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10) + 2
        ws.column_dimensions[col[0].column_letter].width = min(max_len, 45)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"registrations_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# Delete
# ---------------------------------------------------------------------------

@admin_bp.route('/delete/<int:participant_id>', methods=['POST'])
def delete_participant(participant_id: int):
    """Delete a participant record."""
    p = Participant.query.get_or_404(participant_id)
    reg_id = p.registration_id
    db.session.delete(p)
    db.session.commit()
    flash(f'Registration {reg_id} has been deleted.', 'success')
    return redirect(url_for('admin.registrations'))
